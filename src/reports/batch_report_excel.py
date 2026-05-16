from src.data.models.batch import Batch
from openpyxl import Workbook


def generate_batch_report_excel(
        batch: Batch,
        file_path: str
) -> None:
    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = "Batch info"

    info_rows = [
        ("Batch ID", batch.id),
        ("Batch number", batch.batch_number),
        ("Batch date", batch.batch_date.isoformat()),
        ("Status", "Closed" if batch.is_closed else "Open"),
        ("Closed at", batch.closed_at.isoformat() if batch.closed_at else ""),
        ("Work center ID", batch.work_center_id),
        ("Shift", batch.shift),
        ("Team", batch.team),
        ("Task description", batch.task_description),
        ("Nomenclature", batch.nomenclature),
        ("EKN code", batch.ekn_code),
        ("Shift start", batch.shift_start.isoformat()),
        ("Shift end", batch.shift_end.isoformat()),
        ("Created at", batch.created_at.isoformat()),
        ("Updated at", batch.updated_at.isoformat()),
    ]

    for row in info_rows:
        info_sheet.append(row)

    products_sheet = workbook.create_sheet("Products")
    products_sheet.append([
        "ID",
        "Unique code",
        "Aggregated",
        "Aggregated at",
    ])

    for product in batch.products:
        products_sheet.append([
            product.id,
            product.unique_code,
            "Yes" if product.is_aggregated else "No",
            product.aggregated_at.isoformat() if product.aggregated_at else "",
        ])

    stats_sheet = workbook.create_sheet("Statistics")

    total_products = len(batch.products)
    aggregated = sum(1 for product in batch.products if product.is_aggregated)
    remaining = total_products - aggregated
    aggregation_rate = round(aggregated / total_products * 100, 2) if total_products else 0

    stats_rows = [
        ("Total products", total_products),
        ("Aggregated", aggregated),
        ("Remaining", remaining),
        ("Aggregation rate", f"{aggregation_rate}%"),
    ]

    for row in stats_rows:
        stats_sheet.append(row)

    workbook.save(file_path)
